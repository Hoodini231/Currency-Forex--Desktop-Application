import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook as ld
from tkinter import *
from datetime import datetime
import time
import requests
import math
from bs4 import BeautifulSoup
#Required Global Data 
listed= ['Euro', 'Japanese Yen', 'British Pound', 'Australian Dollar', 'Swiss Franc' , 'Chinese Yuan Renminbi'] #Global List of the starting listing currency's based on usage
numbers = []
convertingNumbers = []
currencyNames = []
past_values = []
globalCounting = 0
jo = []
savingList = []
list_of_conversionNumbers = []

#Global Declaration for GUI
root = tk.Tk()
var = listed.copy()

#Declaration for Excel and DATA 
wb2  = Workbook()
ws = wb2.active
ws.title = 'Sheet1'

#Reading DATA into system
wb = ld(filename = 'list.xlsx')
sh = wb['Sheet1']
for i in sh.iter_rows(max_row = 0):
    column = len(i)
    break

for i in range(1,54):
    currencyNames.append(sh.cell(i,1).value)
    past_values.append(sh.cell(i,column).value)



#GUI Creation
root.title('Currency App')
h = 360
w = 640
root.geometry('640x360')
root.iconbitmap('ccIcon.ico')
bgi = tk.PhotoImage(file = 'background.png')

canvas = tk.Canvas(root, width= w, height = h)
canvas.pack(fill = 'both', expand = True)
canvas.create_image(0,0, image=bgi, anchor='nw')
options = currencyNames.copy()
optionsDependent = currencyNames.copy()
optionsDependent.remove('United States Dollar')
'''
def checkFirstLoad():
    with open('DB.text','r') as f:
        lines = f.readlines()
        if lines[0] == 'false/n':
            firstLoad = True
            f.close()
          

def createVar():
    for i in range(6):
        t = StringVar()
        t.set(listed[i])
        var.append(t)
'''

#
# FUNCTIONS RELATED TO GUI
#

def defocus(event):
    event.widget.master.focus_set()

#Drop Text to indicate choices of listed currencies
def drop_windows(list):
    count = 0
    
    for j in range(25,320,294):
        for i in range(100,330,80):
            drop = ttk.Combobox(root,font = 'Helvetica 11', width = 14, value = currencyNames, state = 'readonly')
            drop_window = canvas.create_window(j,i,anchor='nw', window=drop)
            index = currencyNames.index(listed[count])
            drop.current(index)
            drop.bind('<FocusIn>', defocus)
            jo.append(drop)
            jo[count].bind('<<ComboboxSelected>>', newSelection)#So that bindings dont get confused in system
            count = count +1
            '''
            drop = OptionMenu(root, var[count], *list)
            drop.config(width=len("United States  "))
            drop_window = canvas.create_window(j,i,anchor='nw', window=drop, command = selfChanger)
            count = count + 1
            '''
def newSelection(event):
    for i in range(0,5):
        var[i] = jo[i].get()
    changeFunction(False)
    textBoards()

#Declaration for displaying the text labels, for currencies convertion        
def textBoards():
    count = 0
    for j in range(162,460,294):
        for i in range(100,330,80):
            name = var[count]
            index = currencyNames.index(name)
            number = convertingNumbers[index]
            label = tk.Label(root, text = number, font= ('Helvetica', 15),fg = "white", bg = '#1c1c1c')
            label.place(x=j,y=i,relheight=0.09,relwidth =0.15)
            count = count + 1

    info_label = tk.Label(root, text = 'Updating every: 10 seconds', font = 'Helvetica 10', fg ='grey', bg = '#1c1c1c')
    info_label.place(x=465, y=3)

#Declaration for displaying change symbol and amount since last data entry
def changeFunction(boole):
    isRepTrue = False
    bool = boole #Boolean to indicate wheter or not USD 
    #Importation of images neutra, growth and debase?
    neutral = tk.PhotoImage(file = 'neutral.png')
    label2 = tk.Label(canvas, image=neutral)
    label2.image = neutral
    red_Arrow = tk.PhotoImage(file = 'newRed.png')
    labelTry = tk.Label(canvas, image=red_Arrow,)
    labelTry.image = red_Arrow
    green_Arrow = tk.PhotoImage(file = "newGreen.png")
    labelTry2 = tk.Label(canvas, image= green_Arrow)
    labelTry2.image = green_Arrow 

    arrows = [red_Arrow, green_Arrow, neutral]

    updateAllValues()#Update all values within System Before displaying change function.

    if(bool == True):
        conversion_To_New()#Convert listed currencies convesion rate to new one
        
        if(isRepTrue == False): #Checks if its the repeated same currencie         CHECK ON THIS LOGIC
            convert_PastValues()
            isRepTrue = True
        

    count = 0
    intermediate = 0 #Using this variable to control the index's for images
    for j in range(256,650,294):
        for i in range(101,330,80):
            name = var[count]
            change = checkChange(name, convertingNumbers)
            change_visual = round(change, 4)
            colour_indicators = ['red', 'green', 'grey']
            #Conditionals, to determine image usage
            if(change > 0):
                intermediate = 1
            elif(change == 0):
                intermediate = 2
            elif(change < 0):
                intermediate == 0

            label = tk.Label(canvas, image=arrows[intermediate], bg = '#1c1c1c')
            label.place(x=j,y=i,relheight =0.05, relwidth = 0.025, anchor= 'nw')

            value_changed_label = tk.Label(canvas, text = change_visual, bg='#1c1c1c', font = 'Helvetica 7', fg = colour_indicators[intermediate])
            value_changed_label.place(x=j+16,y=i,relheight = 0.05, relwidth = 0.06)
            count = count + 1


'''
def dateTime_Label( strin):
    string = str(strin)
    now = datetime.now()
    currentTime = now.strftime('%H:%M:%S')
    date_label = tk.Label(root, text = 'Last Updated: ' + string + ' seconds ago', font = 'Helvetica 10', fg ='grey', bg = '#1c1c1c')
    date_label.place(x=415, y=3)
'''

def checkChange(name, numbers):
    num = numbers
    cName = name
    index = currencyNames.index(cName)
    past_value = past_values[index]
    change = num[index] - past_value
    return change
'''     
def selected_home(event):
    optionDynamic = options.copy()
    optionDynamic.remove(clicked.get())
    for i in range(0,6):
        if var[i].get() == clicked.get():
            var[i].set('Not Selected')
    drop_windows(optionDynamic)
    home_currency = clicked.get()

a graphing / history of currencies

def newWindow():
    newWindow = tk.Toplevel(root)
    newWindow.geometry('500x300')
    newWindow.title('History of Currency')
    canvas2 = tk.Canvas(root, w= 500, h = 300, bg = 'gray')
    choice = ttk.Combobox(canvas2, values = currencyNames )
    choice.bind("<FocusIn>", defocus)
    choice.bind('<<ComboboxSelected>>', showHistory)

def showHistory(event):
    name = choice.get() 




def more():
    more_image = tk.PhotoImage(file = 'newMore.png')
    more_button = tk.Button(canvas, image = more_image, bg = "#1c1c1c", command = newWindow)
    
    more_button.image = more_image
    more_button.place(x=150,y=3, height = 31, width = 32)

'''

#
#FUNCTIONS RELATED TO LOGICAL FOUNDATION OF APP
#



def updateAllValues():
        convertingNumbers.clear()
        result = requests.get('https://www.x-rates.com/table/?from=USD&amount=1')
        src = result.content
        soup = BeautifulSoup(src ,'html.parser')
        x = soup.find_all('table')
        main = x[1]
        #Scrapping per tags
        for tr_tag in main.find_all('tr'):
            try:
                number = tr_tag.find_all('td')[1].string
                a = str(number)
                b = float(a)
                numbers.append(round(b,3)) 
                #When scrapping, some error as 3 random instances are not in alphabet order, thus the numbers are also not in order             
            except:
               pass
        
        #Correcting the errors
        incoherent_numbers = [numbers[51], numbers[50], numbers[24]]#Grabbing the required numbers for append in order
        internal_clock = 0
        for i in range(0,53): #For Loop to sort out a correcy parallel list 
            if(i == 24 or i == 51 or i == 50):
                pass
            elif(i == 5 or i == 14 or i == 42):
                convertingNumbers.append(incoherent_numbers[internal_clock])
                convertingNumbers.append(numbers[i])
                internal_clock = internal_clock + 1
            else:
                convertingNumbers.append(numbers[i])
        savingList = convertingNumbers.copy() #This ensures that save files are pure to 1USD
        

def cloneList(list): #Clones the old workbook, and adds the new shit
    list_of_conversionNumbers = list 
    print('Start')
    print(list_of_conversionNumbers) 
    #Finding ranges of workbook loaded    
    column = sh['A']
    column_length = len(column) #Y range length
    
    for i in sh.iter_rows(max_row = 0):
        row_length = len(i) # X range length
        break

    #Create cells in new workbook
    for j in range(1,column_length): 
        for i in range(1,row_length):
            ws.cell(row = j, column = i)

    #Cloning workbook before new addition
    for i in range(1,row_length+1):
        for j in range(1,column_length+1):
            ws.cell(row = j, column = i).value = sh.cell(row = j, column = i).value
    save_file('listBackup') #Saved as backup file

    #Adding new listing
    count = 0
    for i in range(1,column_length+1):
        if(i == 53):
            ws.cell(row = i, column = row_length+1).value = 1
        else:
           
            ws.cell(row= i, column = row_length+1).value = list_of_conversionNumbers[count]
            list_of_conversionNumbers.append(ws.cell(row= i, column = row_length+1).value)
            count = count+1
    save_file('list')
    print("File Saved.")

def save_file(name):
    wb2.save(name+'.xlsx')

#Issue, dynamic change,
def conversion_To_New(): #To get past values, we should convert past values as well...
    b = homeCurrencyMenu.get()
    a = convertingNumbers.copy() #Call upon list, once cleared
    convertingNumbers.clear() #To avoid local variable issues, avoid directly equating list
    index = currencyNames.index(b)

    for i in range(0, len(a)):
        if(i == index):
            pass
        elif(i == len(a)-1):
            convertingNumbers.append(a[index])
        else:
            t = round(a[i] / a[index],3 )
            convertingNumbers.append(t)


def convert_PastValues(): #Function for converting past values into past values with value in new home currency
    temp_pastValues = past_values.copy()
    past_values.clear()
    c = homeCurrencyMenu.get()
    index = currencyNames.index(c)
    for i in range(0, len(temp_pastValues)):
        if(i == index):
            pass

        else:
            t = round(temp_pastValues[i] / temp_pastValues[index], 3)
            past_values.append(t)




def save():
    cloneList(convertingNumbers)
      



def Auto_Refresh_Logistics():
    past_values = savingList.copy()



   
'''
def dynamicTimer():
    dateTime_Label(0)
    globalCounting = globalCounting + 1
    root.after(10000, dynamicTimer)
'''
        
class dynamicTime():
    count = 0
    bool = False
    def __init__(self):
        pass

    def action(self):
        
        Auto_Refresh_Logistics()
        changeFunction(self.bool)
        textBoards()
        #dateTime_Label(self.count)
        #print(self.count)
        #self.count = self.count + 1
        root.after(10000, self.action)

    def saveList(self):
        cloneList(convertingNumbers.copy())#This clones over the new numbers scrapped
        root.after(300000, self.saveList)





def test(event):
    home_currency_current = homeCurrencyMenu.get()
    timeChanger.bool = True

def rework(list):
    convertingNumbers = list.copy()
    print(convertingNumbers)


#Runtime main code Start
timeChanger = dynamicTime()
timeChanger.action()


#Creation of more options

#Instantiation of contrast dropbox, number text boards and indicators
drop_windows(optionsDependent)
changeFunction(False)
textBoards()

homeCurrencyMenu = ttk.Combobox(values = options, width = 18, font = 'Helvetica 10')
HCM_window = canvas.create_window(1,3,anchor = 'nw', window=homeCurrencyMenu)
homeCurrencyMenu.current(currencyNames.index('United States Dollar'))
homeCurrencyMenu.bind('<FocusIn>', defocus)
homeCurrencyMenu.bind('<<ComboboxSelected>>', test)
timeChanger.saveList()

root.mainloop()
